#!/usr/bin/env python3
"""
Build field-level shift-share instrument Z_{f,t} = Σ_j Share_{j,f} × VisaRate_{j,t}

Data sources:
  1. CBS InternationalStudents (field × herkomst, 2011/12-2025/26)
  2. CBS EthnicalMakeup (100 nationalities, study permits, 2010-2024)
  3. IND Consular Visa Data (5 Excel files, 2020-2024)

Outputs:
  - instrument_values.csv        (field × year panel)
  - cbs_field_shares.csv         (ω_{j,f} matrix)
  - herkomst_visa_rates.csv      (GroupVisaRate_{j,t})
  - country_crosswalk.csv        (country → herkomst group mapping)
"""

import csv
import glob
import os
import re
import warnings
from collections import defaultdict

import numpy as np
import openpyxl

warnings.filterwarnings("ignore")

DATA_DIR = "/desktop/Thesis work/Data"
OUT_DIR = "/workspace"

# ─── 1. READ CBS INTERNATIONAL STUDENTS ───────────────────────────────────────

def read_international_students():
    """
    Returns a dict: (herkomstland, studierichting, year_int) -> students
    Only 'Totaal mannen en vrouwen' rows (pooled across gender).
    Years are converted to integers: 2011/'12 -> 2011.
    """
    path = os.path.join(DATA_DIR, "CBS Incoming Student Ethnical Makeup",
                        "Ho__eerste__en_ouderejaarsstudenten_19052026_235017.csv")
    data = {}
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            geslacht = row["Geslacht"].strip()
            if geslacht != "Totaal mannen en vrouwen":
                continue
            herkomst = row["Herkomstland"].strip()
            richting = row["Studierichting"].strip()
            period = row["Perioden"].strip()
            val_str = row["Totaal ingeschrevenen (aantal)"].strip()
            # Parse year
            m = re.match(r"(\d{4})", period)
            if not m:
                continue
            year = int(m.group(1))
            # Parse value
            if val_str == "" or val_str == ".":
                continue
            try:
                val = int(val_str.replace(",", ""))
            except ValueError:
                continue
            data[(herkomst, richting, year)] = val
    return data

def get_unique_fields(data):
    """Get sorted unique field names from the data."""
    fields = set()
    for (_, richting, _) in data:
        fields.add(richting)
    return sorted(fields)

def get_unique_herkomst(data):
    """Get sorted unique herkomst groups (excl. Nederland)."""
    herk = set()
    for (h, _, _) in data:
        herk.add(h)
    return sorted(herk)

def compute_field_shares(data, baseline_years=range(2011, 2020)):
    """
    Compute baseline shares ω_{j,f} for each herkomst group j in each field f.
    Uses pre-period years (2011-2019) pooled across education types.
    Returns dict: (herkomstland, richting) -> share
    Also returns total_students_per_field for diagnostics.
    """
    # Aggregate students across baseline years
    agg = defaultdict(int)  # (herkomst, richting) -> total
    field_total = defaultdict(int)  # richting -> total (all herkomst)

    for (h, r, y), v in data.items():
        if y in baseline_years:
            agg[(h, r)] += v
            field_total[r] += v

    # Compute shares for non-Dutch groups
    shares = {}
    for (h, r), total_h_r in agg.items():
        if h == "Nederland":
            continue
        ft = field_total[r]
        if ft > 0:
            shares[(h, r)] = total_h_r / ft

    return shares, agg, field_total

def compute_total_herkomst_shares(data, baseline_years=range(2011, 2020)):
    """
    Compute total shares of each herkomst group across ALL fields (for aggregate instrument).
    Returns dict: herkomstland -> share
    """
    total_non_dutch = 0
    herk_totals = defaultdict(int)
    for (h, r, y), v in data.items():
        if y in baseline_years and h != "Nederland":
            herk_totals[h] += v
            total_non_dutch += v

    shares = {}
    for h, v in herk_totals.items():
        if total_non_dutch > 0:
            shares[h] = v / total_non_dutch
    return shares


# ─── 2. READ CBS ETHNICAL MAKEUP (STUDY PERMITS) ─────────────────────────────

def read_ethnical_makeup():
    """
    Returns dict: (nationality_clean, year_int) -> permits
    Also returns a set of all unique nationalities (stripped).
    """
    path = os.path.join(DATA_DIR, "CBS Incoming Student Ethnical Makeup",
                        "Verblijfsvergunningen_voor_bepaalde_tijd_19052026_175309.csv")
    data = {}
    all_nats = set()
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            nat = row["Nationaliteit"].strip()
            period = row["Perioden"].strip()
            val_str = row["Verblijfsvergunning regulier naar motief/Studie (aantal)"].strip()
            m = re.match(r"(\d{4})", period)
            if not m:
                continue
            year = int(m.group(1))
            try:
                val = int(val_str.replace(",", ""))
            except ValueError:
                continue
            data[(nat, year)] = val
            all_nats.add(nat)
    return data, sorted(all_nats)


# ─── 3. BUILD HERKOMST GROUP → NATIONALITY MAPPING ────────────────────────────

def build_nationality_to_herkomst_map():
    """
    Map the 100 CBS Dutch nationality adjectives to the 7 non-Dutch herkomst groups.
    """
    # Europa (exclusief Nederland) - European countries
    european = [
        "Albanees", "Belg", "Bosnisch", "Brits", "Bulgaars", "Burger Servi\u00eb",
        "Deens", "Duits", "Estisch", "Fins", "Frans", "Grieks",
        "Hongaars", "Iers", "IJslands", "Italiaans", "Kosovaars", "Kroatisch",
        "Lets", "Liechtensteins", "Litouws", "Luxemburgs", "Macedonisch",
        "Maltees", "Moldavisch", "Monaco", "Montenegrijns",
        "Noors", "Oekra\u00efens", "Oostenrijks", "Pools", "Portugees",
        "Roemeens", "Russisch", "San Marinees", "Servisch",
        "Sloveens", "Slowaaks", "Spaans", "Tsjechisch",
        "Vaticaanstad", "Wit-Russisch", "Zweeds", "Zwitsers",
    ]
    # Turkije
    turkish = ["Turks"]
    # Marokko
    moroccan = ["Marokkaans"]
    # Suriname
    surinamese = ["Surinaams"]
    # Nederlandse Cariben
    caribbean = []  # No exact matches in the 100 nationalities list
    # Indonesië
    indonesian = ["Indonesisch"]
    # Buiten-Europa (excl. 5 grote groepen) - everything else
    other = [
        "Afghaans", "Algerijns", "Amerikaans", "Angolees", "Argentijns",
        "Armeens", "Australisch", "Azerbeidzjaans", "Bengalees", "Bhutaans",
        "Boliviaans", "Braziliaans", "Burundisch", "Canadees", "Chileens",
        "Chinees", "Colombiaans", "Congolees", "Congolees (Democratische Republiek)",
        "Costa Ricaans", "Cubaans", "Dominicaans", "Ecuadoraans", "Egyptisch",
        "Eritrees", "Ethiopisch", "Filippijns", "Gambiaans", "Georgisch",
        "Ghanees", "Guinees", "Ha\u00eftiaans", "Indiaas", "Iraaks",
        "Iraans", "Isra\u00eblisch", "Ivoriaans", "Japans", "Jemenitisch",
        "Jordaans", "Kaapverdisch", "Kameroens", "Kazaks", "Keniaans",
        "Koeweits", "Libanees", "Liberiaans", "Libisch", "Maledivisch",
        "Maleisisch", "Mexicaans", "Mongools", "Myanmarees", "Namibisch",
        "Nepalees", "Nieuw-Zeelands", "Nigeriaans", "Oezbeeks", "Omanitisch",
        "Pakistaans", "Peruaans", "Rwandees", "Salvadoraans", "Saoedi-Arabisch",
        "Senegalees", "Sierra Leoons", "Singaporees", "Soedanees", "Somalisch",
        "Sri Lankaans", "Syrisch", "Taiwanees", "Tanzaniaans", "Thais",
        "Togolees", "Tunesisch", "Ugandees", "Uruguayaans", "Venezolaans",
        "Vietnamees", "Zambiaans", "Zimbabwaans", "Zuid-Afrikaans",
        "Zuid-Koreaans",
    ]

    mapping = {}
    for n in european:
        mapping[n] = "Europa (exclusief Nederland)"
    for n in turkish:
        mapping[n] = "Turkije"
    for n in moroccan:
        mapping[n] = "Marokko"
    for n in surinamese:
        mapping[n] = "Suriname"
    for n in caribbean:
        mapping[n] = "Nederlandse Cariben"
    for n in indonesian:
        mapping[n] = "Indonesi\u00eb"
    for n in other:
        mapping[n] = "Buiten-Europa (excl. 5 grote herkomstgroepen)"

    return mapping


# ─── 4. READ IND VISA DATA ────────────────────────────────────────────────────

def read_ind_visa_data():
    """
    Read all 5 IND Excel files. Returns dict: (country_name_upper, year_int) -> acceptance_rate
    Where acceptance_rate = 1 - not_issued_rate_for_uniform_visas.
    Also returns set of all unique IND country names.
    """
    visa_dir = os.path.join(DATA_DIR, "IND Short-Term Schengen Visa Acceptance Rates")
    files = sorted(glob.glob(os.path.join(visa_dir, "*.xlsx*")))

    # Map file names to years
    year_keywords = {
        2020: "2020",
        2021: "2021",
        2022: "2022",
        2023: "2023",
        2024: "2024",
    }

    all_data = {}  # (country, year) -> acceptance_rate
    all_countries = set()

    for fpath in files:
        fname = os.path.basename(fpath)
        # Determine year
        year = None
        for y, kw in year_keywords.items():
            if kw in fname:
                year = y
                break
        if year is None:
            print(f"  WARNING: Could not determine year for {fname}, skipping")
            continue

        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        ws = wb["Data for consulates"]

        # Find column indices from header row
        headers = [str(c).strip() if c else "" for c in next(ws.iter_rows(max_row=1, values_only=True))]
        # Find relevant columns
        country_col = None
        rate_col = None
        for i, h in enumerate(headers):
            if h and "country" in h.lower() and "consulate" in h.lower():
                country_col = i
            if h and "not issued rate" in h.lower() and "uniform" in h.lower():
                rate_col = i

        if country_col is None or rate_col is None:
            print(f"  WARNING: Could not find columns in {fname}")
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= max(country_col, rate_col):
                continue
            country = str(row[country_col]).strip() if row[country_col] else ""
            if not country or country == "None":
                continue

            rate = row[rate_col]
            if rate is None:
                continue
            try:
                rate = float(rate)
            except (ValueError, TypeError):
                continue

            # Calculate acceptance rate = 1 - rejection rate
            acceptance = 1.0 - rate
            all_data[(country, year)] = acceptance
            all_countries.add(country)

        wb.close()

    return all_data, sorted(all_countries)


# ─── 5. BUILD IND COUNTRY → HERKOMST GROUP CROSSWALK ──────────────────────────

def build_ind_to_herkomst_crosswalk(ind_countries, nat_to_herk):
    """
    Map IND English country names to CBS nationality adjectives, then to herkomst groups.
    Returns dict: ind_country -> herkomst_group
    Also returns dict: ind_country -> nationality_adjective (for diagnostics)
    """
    # Manual mapping from IND country names (uppercase) to CBS nationality adjectives
    # This is the key crosswalk
    country_to_nationality = {
        # Europe
        "ALBANIA": "Albanees",
        "ANDORRA": None,  # no permit data
        "AUSTRIA": "Oostenrijks",
        "BELARUS": "Wit-Russisch",
        "BELGIUM": "Belg",
        "BOSNIA AND HERZEGOVINA": "Bosnisch",
        "BULGARIA": "Bulgaars",
        "CROATIA": "Kroatisch",
        "CYPRUS": None,
        "CZECH REPUBLIC": "Tsjechisch",
        "DENMARK": "Deens",
        "ESTONIA": "Estisch",
        "FINLAND": "Fins",
        "FRANCE": "Frans",
        "GERMANY": "Duits",
        "GREECE": "Grieks",
        "HOLY SEE (VATICAN CITY STATE)": "Vaticaanstad",
        "HUNGARY": "Hongaars",
        "ICELAND": "IJslands",
        "IRELAND": "Iers",
        "ITALY": "Italiaans",
        "KOSOVO": "Kosovaars",
        "LATVIA": "Lets",
        "LIECHTENSTEIN": "Liechtensteins",
        "LITHUANIA": "Litouws",
        "LUXEMBOURG": "Luxemburgs",
        "MALTA": "Maltees",
        "MOLDOVA": "Moldavisch",
        "MONACO": "Monaco",
        "MONTENEGRO": "Montenegrijns",
        "NETHERLANDS": None,
        "NORTH MACEDONIA": "Macedonisch",
        "NORWAY": "Noors",
        "POLAND": "Pools",
        "PORTUGAL": "Portugees",
        "ROMANIA": "Roemeens",
        "RUSSIA": "Russisch",
        "RUSSIAN FEDERATION": "Russisch",
        "SAN MARINO": "San Marinees",
        "SERBIA": "Burger Servi\u00eb",
        "SLOVAKIA": "Slowaaks",
        "SLOVENIA": "Sloveens",
        "SPAIN": "Spaans",
        "SWEDEN": "Zweeds",
        "SWITZERLAND": "Zwitsers",
        "TÜRKIYE": "Turks",
        "TURKEY": "Turks",
        "UKRAINE": "Oekra\u00efens",
        "UNITED KINGDOM": "Brits",

        # North America
        "CANADA": "Canadees",
        "UNITED STATES": "Amerikaans",
        "MEXICO": "Mexicaans",

        # Central America & Caribbean
        "ANTIGUA AND BARBUDA": None,
        "BAHAMAS": None,
        "BARBADOS": None,
        "BELIZE": None,
        "COSTA RICA": "Costa Ricaans",
        "CUBA": "Cubaans",
        "DOMINICA": None,
        "DOMINICAN REPUBLIC": "Dominicaans",
        "EL SALVADOR": "Salvadoraans",
        "GRENADA": None,
        "GUATEMALA": None,
        "HAITI": "Ha\u00eftiaans",
        "HONDURAS": None,
        "JAMAICA": None,
        "NICARAGUA": None,
        "PANAMA": None,
        "SAINT KITTS AND NEVIS": None,
        "SAINT LUCIA": None,
        "SAINT VINCENT AND THE GRENADINES": None,
        "TRINIDAD AND TOBAGO": None,

        # South America
        "ARGENTINA": "Argentijns",
        "BOLIVIA": "Boliviaans",
        "BRAZIL": "Braziliaans",
        "CHILE": "Chileens",
        "COLOMBIA": "Colombiaans",
        "ECUADOR": "Ecuadoraans",
        "GUYANA": None,
        "PARAGUAY": None,
        "PERU": "Peruaans",
        "SURINAME": "Surinaams",
        "URUGUAY": "Uruguayaans",
        "VENEZUELA": "Venezolaans",

        # Africa
        "ALGERIA": "Algerijns",
        "ANGOLA": "Angolees",
        "BENIN": None,
        "BOTSWANA": None,
        "BURKINA FASO": None,
        "BURUNDI": "Burundisch",
        "CAPE VERDE": "Kaapverdisch",
        "CABO VERDE": "Kaapverdisch",
        "CAMEROON": "Kameroens",
        "CENTRAL AFRICAN REPUBLIC": None,
        "CHAD": None,
        "COMOROS": None,
        "CONGO": "Congolees",
        "DEMOCRATIC REPUBLIC OF THE CONGO": "Congolees (Democratische Republiek)",
        "DJIBOUTI": None,
        "EGYPT": "Egyptisch",
        "EQUATORIAL GUINEA": None,
        "ERITREA": "Eritrees",
        "ETHIOPIA": "Ethiopisch",
        "GABON": None,
        "GAMBIA": "Gambiaans",
        "GHANA": "Ghanees",
        "GUINEA": "Guinees",
        "GUINEA-BISSAU": None,
        "IVORY COAST": "Ivoriaans",
        "KENYA": "Keniaans",
        "LESOTHO": None,
        "LIBERIA": "Liberiaans",
        "LIBYA": "Libisch",
        "MADAGASCAR": None,
        "MALAWI": None,
        "MALI": None,
        "MAURITANIA": None,
        "MAURITIUS": None,
        "MOROCCO": "Marokkaans",
        "MOZAMBIQUE": None,
        "NAMIBIA": "Namibisch",
        "NIGER": None,
        "NIGERIA": "Nigeriaans",
        "RWANDA": "Rwandees",
        "SAO TOME AND PRINCIPE": None,
        "SENEGAL": "Senegalees",
        "SEYCHELLES": None,
        "SIERRA LEONE": "Sierra Leoons",
        "SOMALIA": "Somalisch",
        "SOUTH AFRICA": "Zuid-Afrikaans",
        "SOUTH SUDAN": None,
        "SUDAN": "Soedanees",
        "ESWATINI": None,
        "TANZANIA": "Tanzaniaans",
        "TOGO": "Togolees",
        "TUNISIA": "Tunesisch",
        "UGANDA": "Ugandees",
        "ZAMBIA": "Zambiaans",
        "ZIMBABWE": "Zimbabwaans",

        # Asia
        "AFGHANISTAN": "Afghaans",
        "ARMENIA": "Armeens",
        "AZERBAIJAN": "Azerbeidzjaans",
        "BAHRAIN": None,
        "BANGLADESH": "Bengalees",
        "BHUTAN": "Bhutaans",
        "BRUNEI": None,
        "CAMBODIA": None,
        "CHINA": "Chinees",
        "EAST TIMOR": None,
        "GEORGIA": "Georgisch",
        "INDIA": "Indiaas",
        "INDONESIA": "Indonesisch",
        "IRAN": "Iraans",
        "IRAQ": "Iraaks",
        "ISRAEL": "Isra\u00eblisch",
        "JAPAN": "Japans",
        "JORDAN": "Jordaans",
        "KAZAKHSTAN": "Kazaks",
        "KUWAIT": "Koeweits",
        "KYRGYZSTAN": None,
        "LAOS": None,
        "LEBANON": "Libanees",
        "MALAYSIA": "Maleisisch",
        "MALDIVES": "Maledivisch",
        "MONGOLIA": "Mongools",
        "MYANMAR": "Myanmarees",
        "NEPAL": "Nepalees",
        "NORTH KOREA": None,
        "OMAN": "Omanitisch",
        "PAKISTAN": "Pakistaans",
        "PALESTINE": None,
        "PHILIPPINES": "Filippijns",
        "QATAR": None,
        "SAUDI ARABIA": "Saoedi-Arabisch",
        "SINGAPORE": "Singaporees",
        "SOUTH KOREA": "Zuid-Koreaans",
        "SRI LANKA": "Sri Lankaans",
        "SYRIA": "Syrisch",
        "TAIWAN": "Taiwanees",
        "TAJIKISTAN": None,
        "THAILAND": "Thais",
        "TURKMENISTAN": None,
        "UNITED ARAB EMIRATES": None,
        "UZBEKISTAN": "Oezbeeks",
        "USA": "Amerikaans",
        "VIET NAM": "Vietnamees",
        "VIETNAM": "Vietnamees",
        "YEMEN": "Jemenitisch",

        # Extra uncommon / alternative spellings
        "COTE D'IVOIRE": "Ivoriaans",
        "COTE IVOIRE": "Ivoriaans",
        "KOREA, REPUBLIC OF": "Zuid-Koreaans",
        "RUMANIA": "Roemeens",
        "FORMER YUGOSLAV REPUBLIC OF MACEDONIA": "Macedonisch",
        "HONG KONG S.A.R.": "Chinees",
        "MACAO S.A.R.": "Chinees",
        "CAPE VERDE": "Kaapverdisch",

        # Oceania
        "AUSTRALIA": "Australisch",
        "FIJI": None,
        "KIRIBATI": None,
        "MARSHALL ISLANDS": None,
        "MICRONESIA": None,
        "NAURU": None,
        "NEW ZEALAND": "Nieuw-Zeelands",
        "PALAU": None,
        "PAPUA NEW GUINEA": None,
        "SAMOA": None,
        "SOLOMON ISLANDS": None,
        "TONGA": None,
        "TUVALU": None,
        "VANUATU": None,
    }

    crosswalk = {}
    nat_mapping = {}
    for ind_country in ind_countries:
        # Try exact match
        if ind_country in country_to_nationality:
            nat_adj = country_to_nationality[ind_country]
        else:
            # Try fuzzy matching
            found = False
            for key, val in country_to_nationality.items():
                if ind_country.startswith(key) or key.startswith(ind_country):
                    nat_adj = val
                    found = True
                    break
            if not found:
                nat_adj = None

        if nat_adj and nat_adj in nat_to_herk:
            crosswalk[ind_country] = nat_to_herk[nat_adj]
            nat_mapping[ind_country] = nat_adj
        else:
            crosswalk[ind_country] = None
            nat_mapping[ind_country] = None

    return crosswalk, nat_mapping


# ─── 6. COMPUTE HERKOMST-GROUP-LEVEL VISA RATES ──────────────────────────────

def compute_group_visa_rates(visa_data, ind_to_herk, permits_data, nat_to_herk,
                             ind_to_nat, visa_years=range(2020, 2025)):
    """
    For each herkomst group j and year t, compute weighted average visa acceptance rate:
    GroupVisaRate_{j,t} = Σ_{c∈j} share_{c,j} × (1 - not_issued_rate_{c,t})
    where shares come from EthnicalMakeup study permits.
    """
    # Step 1: Compute country-level share within each herkomst group using permits
    # Aggregate permits across all years for stable weights
    permits_by_nat = defaultdict(int)  # nationality -> total permits
    for (nat, year), val in permits_data.items():
        permits_by_nat[nat] += val

    # Map nationalities to herkomst groups
    permits_by_herk = defaultdict(int)  # herkomst -> total permits
    for nat, val in permits_by_nat.items():
        h = nat_to_herk.get(nat)
        if h:
            permits_by_herk[h] += val

    # Country shares WITHIN each herkomst group
    nat_shares = {}  # (herkomst, nationality) -> share
    for nat, val in permits_by_nat.items():
        h = nat_to_herk.get(nat)
        if h and permits_by_herk[h] > 0:
            nat_shares[(h, nat)] = val / permits_by_herk[h]

    # Step 2: For each herkomst group and year, compute weighted avg visa rate
    group_rates = {}  # (herkomst, year) -> acceptance_rate

    # First, map IND countries to nationalities and then to herkomst
    ind_nat_year_rates = {}  # (nationality, year) -> list of acceptance rates
    for (ind_country, year), rate in visa_data.items():
        nat = ind_to_nat.get(ind_country)
        if not nat:
            continue
        ind_nat_year_rates.setdefault((nat, year), []).append(rate)

    # Average rates if a nationality appears multiple times in IND data
    ind_nat_year_rate = {}
    for (nat, year), rates in ind_nat_year_rates.items():
        ind_nat_year_rate[(nat, year)] = np.mean(rates)

    # Now compute group-level rates
    for h in set(nat_to_herk.values()):
        for y in visa_years:
            weighted_sum = 0.0
            weight_total = 0.0
            for nat, val in permits_by_nat.items():
                if nat_to_herk.get(nat) != h:
                    continue
                share = nat_shares.get((h, nat), 0)
                rate = ind_nat_year_rate.get((nat, y))
                if rate is not None and share > 0:
                    weighted_sum += share * rate
                    weight_total += share
            if weight_total > 0:
                # Normalize by actual coverage
                group_rates[(h, y)] = weighted_sum / weight_total
            else:
                group_rates[(h, y)] = None

    return group_rates, nat_shares, permits_by_nat, permits_by_herk


# ─── 7. COMPUTE FIELD-LEVEL INSTRUMENT ────────────────────────────────────────

def compute_instrument(field_shares, group_visa_rates, fields, herkomst_groups,
                       visa_years=range(2020, 2025), total_herk_shares=None):
    """
    Compute Z_{f,t} = Σ_{j≠NL} ω_{j,f} × GroupVisaRate_{j,t}
    Also compute Z_t (aggregate): Σ_{j≠NL} total_share_j × GroupVisaRate_{j,t}
    """
    results = []
    for field in fields:
        for year in visa_years:
            z_field = 0.0
            weight_sum = 0.0
            for h in herkomst_groups:
                if h == "Nederland":
                    continue
                share = field_shares.get((h, field), 0)
                rate = group_visa_rates.get((h, year))
                if rate is not None and share > 0:
                    z_field += share * rate
                    weight_sum += share
            if weight_sum > 0:
                # Normalize by sum of available weights (in case some groups missing)
                z_field = z_field / weight_sum if weight_sum > 0 else z_field
            results.append((field, year, z_field))

    # Aggregate instrument (same for all fields, varies by year)
    agg_results = {}
    for year in visa_years:
        z_agg = 0.0
        weight_sum = 0.0
        for h in herkomst_groups:
            if h == "Nederland":
                continue
            share = total_herk_shares.get(h, 0)
            rate = group_visa_rates.get((h, year))
            if rate is not None and share > 0:
                z_agg += share * rate
                weight_sum += share
        if weight_sum > 0:
            z_agg = z_agg / weight_sum
        agg_results[year] = z_agg

    return results, agg_results


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("BUILDING FIELD-LEVEL SHIFT-SHARE INSTRUMENT")
    print("=" * 70)

    # ── Step 1: Read International Students ──
    print("\n[1/7] Reading CBS InternationalStudents...")
    students = read_international_students()
    fields = get_unique_fields(students)
    herkomst_groups = get_unique_herkomst(students)
    print(f"  Fields ({len(fields)}): {fields}")
    print(f"  Herkomst groups ({len(herkomst_groups)}): {herkomst_groups}")
    print(f"  Total observations: {len(students)}")

    # ── Step 2: Compute baseline shares ──
    print("\n[2/7] Computing field-level herkomst shares (baseline: 2011-2019)...")
    field_shares, agg_students, field_totals = compute_field_shares(students)
    total_herk_shares = compute_total_herkomst_shares(students)

    print("  Field shares ω_{j,f}:")
    for (h, r), s in sorted(field_shares.items()):
        if s > 0.01:
            print(f"    {r}: {h} = {s:.4f}")
    print(f"  Total {len(field_shares)} non-zero shares")

    print("\n  Total herkomst shares (across all fields):")
    for h, s in sorted(total_herk_shares.items(), key=lambda x: -x[1]):
        print(f"    {h}: {s:.4f}")

    # ── Step 3: Read EthnicalMakeup ──
    print("\n[3/7] Reading CBS EthnicalMakeup (study permits)...")
    permits_data, all_nationalities = read_ethnical_makeup()
    print(f"  Total nationalities: {len(all_nationalities)}")
    print(f"  Total observations: {len(permits_data)}")

    # ── Step 4: Build nationality → herkomst mapping ──
    print("\n[4/7] Building nationality → herkomst group mapping...")
    nat_to_herk = build_nationality_to_herkomst_map()

    # Check coverage
    mapped = sum(1 for n in all_nationalities if n in nat_to_herk)
    unmapped = [n for n in all_nationalities if n not in nat_to_herk]
    print(f"  Mapped: {mapped}/{len(all_nationalities)} nationalities")
    if unmapped:
        print(f"  Unmapped: {unmapped}")

    # Check distribution
    herk_counts = defaultdict(int)
    for n, h in nat_to_herk.items():
        if n in all_nationalities:
            herk_counts[h] += 1
    for h, c in sorted(herk_counts.items()):
        print(f"    {h}: {c} nationalities")

    # ── Step 5: Read IND visa data ──
    print("\n[5/7] Reading IND consular visa data (2020-2024)...")
    visa_data, ind_countries = read_ind_visa_data()
    print(f"  IND countries: {len(ind_countries)}")
    print(f"  Total observations: {len(visa_data)}")

    # Check year coverage
    visa_years = set()
    for (_, y) in visa_data:
        visa_years.add(y)
    print(f"  Years with data: {sorted(visa_years)}")

    # ── Step 6: Build IND → herkomst crosswalk ──
    print("\n[6/7] Building IND country → herkomst group crosswalk...")
    ind_to_herk, ind_to_nat = build_ind_to_herkomst_crosswalk(ind_countries, nat_to_herk)

    # Diagnostics
    matched = sum(1 for v in ind_to_herk.values() if v is not None)
    unmatched = [c for c, v in ind_to_herk.items() if v is None]
    print(f"  Matched: {matched}/{len(ind_countries)} IND countries")
    if unmatched:
        print(f"  Unmatched IND countries ({len(unmatched)}): {unmatched}")

    # ── Compute herkomst-group visa rates ──
    print("\n  Computing herkomst-group visa acceptance rates...")
    group_rates, nat_shares, permits_by_nat, permits_by_herk = compute_group_visa_rates(
        visa_data, ind_to_herk, permits_data, nat_to_herk, ind_to_nat,
        visa_years=range(2020, 2025)
    )

    print("  GroupVisaRate_{j,t}:")
    for (h, y), r in sorted(group_rates.items()):
        if r is not None:
            print(f"    {h}, {y}: {r:.4f}")
        else:
            print(f"    {h}, {y}: NO DATA")

    # ── Step 7: Compute field-level instrument ──
    print("\n[7/7] Computing field-level instrument Z_{f,t}...")
    instrument_results, agg_results = compute_instrument(
        field_shares, group_rates, fields, herkomst_groups,
        visa_years=range(2020, 2025), total_herk_shares=total_herk_shares
    )

    print("\n  Z_{f,t} (field-varying instrument):")
    for field, year, z in instrument_results:
        print(f"    {field}, {year}: {z:.4f}")

    print("\n  Z_t (aggregate instrument, same for all fields):")
    for year, z in sorted(agg_results.items()):
        print(f"    {year}: {z:.4f}")

    # ── WRITE OUTPUTS ──

    # 1. Instrument values CSV
    out_path = os.path.join(OUT_DIR, "instrument_values.csv")
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["field", "year", "Z_field", "Z_agg"])
        for field, year, z_field in instrument_results:
            z_agg = agg_results.get(year, None)
            writer.writerow([field, year, f"{z_field:.6f}", f"{z_agg:.6f}" if z_agg else ""])
    print(f"\n  ✓ Wrote: {out_path}")

    # 2. Field shares CSV
    out_path2 = os.path.join(OUT_DIR, "cbs_field_shares.csv")
    with open(out_path2, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        header = ["field"] + [h for h in herkomst_groups if h != "Nederland"]
        writer.writerow(header)
        for field in fields:
            row = [field]
            for h in herkomst_groups:
                if h == "Nederland":
                    continue
                row.append(f"{field_shares.get((h, field), 0):.6f}")
            writer.writerow(row)
    print(f"  ✓ Wrote: {out_path2}")

    # 3. Herkomst visa rates CSV
    out_path3 = os.path.join(OUT_DIR, "herkomst_visa_rates.csv")
    with open(out_path3, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["herkomst_group", "year", "visa_acceptance_rate"])
        for (h, y), r in sorted(group_rates.items()):
            writer.writerow([h, y, f"{r:.6f}" if r is not None else ""])
    print(f"  ✓ Wrote: {out_path3}")

    # 4. Country crosswalk CSV
    out_path4 = os.path.join(OUT_DIR, "country_crosswalk.csv")
    with open(out_path4, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["ind_country", "cbs_nationality", "herkomst_group"])
        for ind_c in sorted(ind_countries):
            nat = ind_to_nat.get(ind_c, "")
            herk = ind_to_herk.get(ind_c, "")
            writer.writerow([ind_c, nat if nat else "", herk if herk else ""])
    print(f"  ✓ Wrote: {out_path4}")

    # ── DIAGNOSTIC SUMMARY ──
    print("\n" + "=" * 70)
    print("DIAGNOSTIC SUMMARY")
    print("=" * 70)

    print(f"\nCoverage:")
    print(f"  CBS InternationalStudents: {len(students)} rows")
    print(f"  CBS EthnicalMakeup: {len(all_nationalities)} nationalities")
    print(f"  IND visa countries: {len(ind_countries)}")
    print(f"  Crosswalk match rate: {matched}/{len(ind_countries)} ({100*matched/len(ind_countries):.1f}%)")

    # Permit coverage by herkomst group
    print(f"\n  Study permit distribution by herkomst group:")
    for h, v in sorted(permits_by_herk.items(), key=lambda x: -x[1]):
        print(f"    {h}: {v} total permits")

    # Visa rate coverage
    print(f"\n  Herkomst-group visa rate coverage:")
    for h in sorted(set(nat_to_herk.values())):
        n_years = sum(1 for (gh, gy), r in group_rates.items() if gh == h and r is not None)
        print(f"    {h}: {n_years}/5 years with data")

    # Instrument summary stats
    z_field_vals = [z for _, _, z in instrument_results if z != 0 and not np.isnan(z)]
    if z_field_vals:
        print(f"\n  Z_field summary:")
        print(f"    Mean: {np.mean(z_field_vals):.4f}")
        print(f"    Std:  {np.std(z_field_vals):.4f}")
        print(f"    Min:  {np.min(z_field_vals):.4f}")
        print(f"    Max:  {np.max(z_field_vals):.4f}")
        print(f"    N:    {len(z_field_vals)}")

    z_agg_vals = [z for z in agg_results.values() if z != 0 and not np.isnan(z)]
    if z_agg_vals:
        print(f"\n  Z_agg summary:")
        print(f"    Mean: {np.mean(z_agg_vals):.4f}")
        print(f"    Std:  {np.std(z_agg_vals):.4f}")
        print(f"    Min:  {np.min(z_agg_vals):.4f}")
        print(f"    Max:  {np.max(z_agg_vals):.4f}")
        print(f"    N:    {len(z_agg_vals)}")

    print("\n" + "=" * 70)
    print("DONE")
    print("=" * 70)


if __name__ == "__main__":
    main()
